from openpyxl import load_workbook

class Excel_Functions:
   def __init__(self, excel_file_name, excel_sheet_name):
       self.file = excel_file_name
       self.sheet = excel_sheet_name


   # fetch the row count
   def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


   # fetch the column count
   def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


   # read data from excel file
   def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.cell(row=row_number, column=column_number).value


   # write data into excel file
   def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)

